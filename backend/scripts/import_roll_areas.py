"""
Import roll area information from Excel material names.
Material names contain dimensions like "1,6x100 m" which indicates width x length in meters.
"""
import sys
import os
import re
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from app.db.session import SessionLocal
from app.db.models.material import Material

# Excel file path
EXCEL_FILE = '/Users/josemariabarbeito/PycharmProjects/DeltaDash/Calculadora de Consumos de Chalecos Balistico - DEV-INEX-PROD-SHEET-01.xlsx'

# Load the workbook
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)

# Get all material names from all vest sheets
vest_sheets = [
    'MULTIHIT_II', 'ULTRA STOP TCA II', 'STOP III', 'STOP III_B', 'STOP II',
    'MDS_II', 'MDS_II_LIGHT', 'MDS_III', 'DEF_III', 'ULTRA_STOP_III',
    'STOP_FORCE_RF1', 'Lateral_RF1', 'MISS_III', 'LADY_STOP_III',
    'LADY_STOP_III_B', 'LADY_STOP_III_C', 'LADY_STOP_II', 'MS_II',
    'GEOTEX_COMFORT_III'
]

material_roll_areas = {}

for sheet_name in vest_sheets:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    
    # Parse material names from column A (rows 4-30 typically)
    for row in range(4, 35):
        material_name = ws.cell(row=row, column=1).value
        if material_name and isinstance(material_name, str):
            # Extract dimensions from material name
            # Pattern: "1,6x100 m" or "1.6x100 m" or "1,6x100m"
            match = re.search(r'(\d+[.,]\d+)[xX](\d+)\s*m', material_name)
            if match:
                width = float(match.group(1).replace(',', '.'))
                length = float(match.group(2))
                roll_area = width * length
                
                if material_name not in material_roll_areas:
                    material_roll_areas[material_name] = roll_area
                print(f"Found: {material_name} -> {width}m x {length}m = {roll_area} m²")

print(f"\nTotal materials with roll info: {len(material_roll_areas)}")

# Update materials in database
db = SessionLocal()

try:
    updated_count = 0
    # Create a mapping from Excel names to database names
    name_mapping = {
        'Polietileno UAE UD161 1,6x100 m': 'KL 161',
        'Polietileno UD KL230 1,55x120 m': 'KL 230',
        'Polietileno UD Flex KS3000 1,6x150 m': 'FLEX_KS3000',
        'Aramida UD234 1,6x150 m': 'UD 234',
        'Aramida UD245 1,6x150 m': 'UD 245',
        'Poliester No Tejido Geotextil 500g/m² 1,6x50 m': 'GeoTextil 500',
        'Poliester No Tejido Geotextil Blanco 300g/m² 1,5x100 m': 'Geotextil 300',
        'Poliester no Tejido Geotextil 310 g/m² VL930C 1,5x50 m': 'GeoTexile 310',
        'Nylon 70D 200 g/m² con TPU Negro 1,5x100 m': None,  # May not exist
        'Nylon Rockdura 500D Negro 1,5x100 m': None,  # May not exist
    }
    
    for excel_name, roll_area in material_roll_areas.items():
        db_name = name_mapping.get(excel_name)
        if not db_name:
            continue
        
        material = db.query(Material).filter(Material.name == db_name).first()
        if material and material.roll_area_m2 is None:
            material.roll_area_m2 = roll_area
            print(f"Updated {material.name} with roll area {roll_area} m²")
            updated_count += 1
    
    db.commit()
    print(f"\nSuccessfully updated {updated_count} materials with roll area information!")
    
except Exception as e:
    db.rollback()
    print(f"\nError updating materials: {e}")
    import traceback
    traceback.print_exc()
finally:
    db.close()
