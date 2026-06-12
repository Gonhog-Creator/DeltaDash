"""
Comprehensive script to import all geometries, materials, and configurations from Excel.
This script:
1. Imports geometries and surface areas from the Superficies sheet
2. Creates all materials found in vest sheets
3. Creates geometry material configurations for carrier/accessories
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import re
from app.db.session import SessionLocal
from app.db.models.geometry import Geometry
from app.db.models.geometry_material_config import GeometryMaterialConfig
from app.db.models.material import Material

# Excel file path
EXCEL_FILE = '/Users/josemariabarbeito/PycharmProjects/DeltaDash/Calculadora de Consumos de Chalecos Balistico - DEV-INEX-PROD-SHEET-01.xlsx'

# Load the workbook
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)

# Get database session
db = SessionLocal()

try:
    # ============================================
    # STEP 1: Import Geometries from Superficies sheet
    # ============================================
    print("=" * 60)
    print("STEP 1: Importing Geometries from Superficies sheet")
    print("=" * 60)
    
    if 'Superficies' not in wb.sheetnames:
        print("Superficies sheet not found!")
    else:
        ws = wb['Superficies']
        
        # Parse geometries from Superficies sheet
        # Expected format: Row 1 has headers, Row 2+ has geometry data
        # Columns: Geometry Name, XS Front, XS Back, S Front, S Back, etc.
        
        geometries_created = 0
        geometries_skipped = 0
        
        for row in range(2, 50):
            geometry_name = ws.cell(row=row, column=1).value
            if not geometry_name or not isinstance(geometry_name, str):
                continue
            
            # Skip if already exists
            existing = db.query(Geometry).filter(Geometry.name == geometry_name).first()
            if existing:
                print(f"  Geometry {geometry_name} already exists, skipping")
                geometries_skipped += 1
                continue
            
            # Parse surface areas for each size
            surface_areas = {}
            sizes = ['XS', 'S', 'M', 'L', 'XL', 'XXL']
            
            for idx, size in enumerate(sizes):
                front_col = 2 + (idx * 2)
                back_col = 3 + (idx * 2)
                
                front_area = ws.cell(row=row, column=front_col).value
                back_area = ws.cell(row=row, column=back_col).value
                
                if front_area and isinstance(front_area, (int, float)):
                    surface_areas[size] = {
                        "front": float(front_area),
                        "back": float(back_area) if back_area and isinstance(back_area, (int, float)) else 0.0
                    }
            
            if surface_areas:
                # Determine vest type based on geometry name
                vest_type = 'Soft'
                if 'SAPI' in geometry_name.upper() or 'Lateral' in geometry_name.upper():
                    vest_type = 'Hard'
                
                geometry = Geometry(
                    name=geometry_name,
                    vest_type=vest_type,
                    surface_areas=surface_areas,
                    available_sizes=list(surface_areas.keys())
                )
                db.add(geometry)
                geometries_created += 1
                print(f"  Created geometry: {geometry_name} with {len(surface_areas)} sizes")
        
        db.commit()
        print(f"\nGeometries: {geometries_created} created, {geometries_skipped} skipped")
    
    # ============================================
    # STEP 2: Import Materials from all vest sheets
    # ============================================
    print("\n" + "=" * 60)
    print("STEP 2: Importing Materials from vest sheets")
    print("=" * 60)
    
    # Get all vest sheets (exclude non-vest sheets)
    exclude_sheets = ['CALCULADOR', 'Superficies', 'AUX', 'STOCK']
    vest_sheets = [s for s in wb.sheetnames if s not in exclude_sheets]
    
    materials_created = 0
    materials_skipped = 0
    
    for sheet_name in vest_sheets:
        ws = wb[sheet_name]
        
        for row in range(4, 35):
            material_name = ws.cell(row=row, column=1).value
            if not material_name or not isinstance(material_name, str):
                continue
            
            # Skip if already exists
            existing = db.query(Material).filter(Material.name == material_name).first()
            if existing:
                materials_skipped += 1
                continue
            
            # Determine material class based on name
            material_class = 'ballistic'
            if 'Nylon' in material_name:
                material_class = 'fabric'
            elif 'Velcro' in material_name:
                material_class = 'accessory'
            elif 'Elástico' in material_name:
                material_class = 'accessory'
            elif 'Tela' in material_name or 'Spacer' in material_name:
                material_class = 'fabric'
            elif 'Geotextil' in material_name or 'GeoTextil' in material_name:
                material_class = 'ballistic'
            elif 'Vitelmat' in material_name:
                material_class = 'ballistic'
            elif 'PEAD' in material_name:
                material_class = 'ballistic'
            elif 'Honeycomb' in material_name:
                material_class = 'ballistic'
            elif 'Plate' in material_name or 'Placa' in material_name:
                material_class = 'ballistic'
            
            # Extract areal density if present (e.g., "200 g/m²")
            areal_density = None
            if 'g/m²' in material_name:
                match = re.search(r'(\d+)\s*g/m²', material_name)
                if match:
                    areal_density = float(match.group(1))
            
            # Extract roll dimensions if present (e.g., "1,5x100 m")
            roll_area = None
            if 'x' in material_name and 'm' in material_name:
                match = re.search(r'([\d,]+)\s*x\s*([\d,]+)\s*m', material_name)
                if match:
                    width = float(match.group(1).replace(',', '.'))
                    length = float(match.group(2).replace(',', '.'))
                    roll_area = width * length
            
            material = Material(
                name=material_name,
                material_class=material_class,
                areal_density_g_m2=areal_density,
                roll_area_m2=roll_area
            )
            db.add(material)
            materials_created += 1
            print(f"  Created material: {material_name} ({material_class})")
    
    db.commit()
    print(f"\nMaterials: {materials_created} created, {materials_skipped} skipped")
    
    # ============================================
    # STEP 3: Import Geometry Material Configurations
    # ============================================
    print("\n" + "=" * 60)
    print("STEP 3: Importing Geometry Material Configurations")
    print("=" * 60)
    
    # Mapping from vest sheet names to geometry names
    # This needs to be updated based on actual Excel structure
    vest_to_geometry = {
        'ULTRA STOP TCA II': 'DELTA II',
        'STOP III': 'DELTA III',
        'STOP III_B': 'DELTA III',
        'STOP II': 'DELTA I',
        'MDS_II': 'DELTA II',
        'MDS_II_LIGHT': 'DELTA II',
        'MDS_III': 'DELTA III',
        'DEF_III': 'DELTA III',
        'ULTRA_STOP_III': 'DELTA III',
        'MISS_III': 'DELTA III',
        'LADY_STOP_III': 'DELTA III',
        'LADY_STOP_III_B': 'DELTA III',
        'LADY_STOP_III_C': 'DELTA III',
        'LADY_STOP_II': 'DELTA I',
        'MS_II': 'DELTA II',
        'GEOTEX_COMFORT_III': 'DELTA III',
        'MULTIHIT_II': 'DELTA II',
        'STOP_FORCE_RF1': 'DELTA II',
        'Lateral_RF1': 'Lateral',
    }
    
    configs_created = 0
    configs_skipped = 0
    
    for sheet_name, geometry_name in vest_to_geometry.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found, skipping")
            configs_skipped += 1
            continue
        
        ws = wb[sheet_name]
        
        # Find the geometry
        geometry = db.query(Geometry).filter(Geometry.name == geometry_name).first()
        if not geometry:
            print(f"Geometry {geometry_name} not found in database, skipping")
            configs_skipped += 1
            continue
        
        print(f"\nProcessing {geometry_name} from sheet {sheet_name}")
        
        # Parse accessories (carrier materials only)
        accessories = []
        
        for row in range(4, 35):
            material_name = ws.cell(row=row, column=1).value
            layer_count = ws.cell(row=row, column=3).value
            part_number = ws.cell(row=row, column=2).value
            
            if not material_name or not isinstance(material_name, str):
                continue
            
            # Check if this is an outer cover/accessory material
            is_outer_cover = any(keyword in material_name for keyword in ['Nylon', 'Velcro', 'Elástico', 'Tela', 'Spacer'])
            
            if is_outer_cover:
                # Find the material
                material = db.query(Material).filter(Material.name == material_name).first()
                if not material:
                    print(f"  Material {material_name} not found, skipping")
                    continue
                
                # Get quantity from size columns
                quantity = ws.cell(row=row, column=5).value
                
                # If column 5 is a formula, try to use layer count
                if quantity and isinstance(quantity, str) and quantity.startswith('='):
                    quantity = layer_count if layer_count and isinstance(layer_count, (int, float)) else None
                
                if quantity and isinstance(quantity, (int, float)):
                    accessories.append({
                        "material_id": str(material.id),
                        "quantity_per_vest": float(quantity),
                        "unit": "meters",
                        "notes": f"Part: {part_number if part_number else 'N/A'}"
                    })
                    print(f"  Added accessory: {material.name} - {quantity}m per vest")
        
        # Create configuration for "ALL" sizes
        if accessories:
            # Check if config already exists
            existing = db.query(GeometryMaterialConfig).filter(
                GeometryMaterialConfig.geometry_id == geometry.id,
                GeometryMaterialConfig.size == "ALL"
            ).first()
            
            if existing:
                print(f"  Config already exists for {geometry_name} - ALL, skipping")
                configs_skipped += 1
            else:
                config = GeometryMaterialConfig(
                    geometry_id=geometry.id,
                    size="ALL",
                    material_requirements=[],  # Empty - ballistic materials are in Vest → VestLayer
                    accessories=accessories,
                    efficiency_factor=1.15,
                    notes=f"Imported from Excel sheet {sheet_name}"
                )
                db.add(config)
                configs_created += 1
                print(f"  Created config for {geometry_name} - ALL with {len(accessories)} accessories")
        else:
            print(f"  No accessories found for {geometry_name}")
            configs_skipped += 1
    
    db.commit()
    print(f"\nConfigurations: {configs_created} created, {configs_skipped} skipped")
    
    print("\n" + "=" * 60)
    print("IMPORT COMPLETE")
    print("=" * 60)
    print(f"Geometries: {geometries_created} created, {geometries_skipped} skipped")
    print(f"Materials: {materials_created} created, {materials_skipped} skipped")
    print(f"Configurations: {configs_created} created, {configs_skipped} skipped")
    
except Exception as e:
    db.rollback()
    print(f"\nError during import: {e}")
    import traceback
    traceback.print_exc()
finally:
    db.close()
