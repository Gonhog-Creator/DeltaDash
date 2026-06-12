"""
Import geometry data from Excel Superficies sheet into the database.
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from app.db.session import SessionLocal
from app.db.models.geometry import Geometry

# Excel file path
EXCEL_FILE = '/Users/josemariabarbeito/PycharmProjects/DeltaDash/Calculadora de Consumos de Chalecos Balistico - DEV-INEX-PROD-SHEET-01.xlsx'

# Load the workbook without data_only to get formulas, then evaluate them
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
ws = wb['Superficies']

# Parse the data
# Structure from Excel:
# Row 2: Headers with geometry names (DELTA II, DELTA I, DELTA III, FEM II, FEM III, SAPI, Lateral, DELTA IV)
# Row 3: Sub-headers (Frente, Espalda, TOTAL repeated for each geometry)
# Row 4+: Data rows with sizes and surface areas

# Get geometry names from row 1 (columns 4, 8, 12, 16, 20, 24, 28, 32)
geometry_names = []
geometry_columns = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=1, column=col).value
    if cell_value and cell_value not in geometry_names:
        geometry_names.append(cell_value)
        geometry_columns.append(col)

print(f"Found geometries: {geometry_names}")

# Parse surface areas for each geometry
geometries_data = {}

for row in range(4, ws.max_row):
    size = ws.cell(row=row, column=2).value
    if not size or size not in ['XS', 'S', 'M', 'L', 'XL', 'XXL']:
        continue
    
    # For each geometry, get front and back areas using the column positions
    for idx, geom_name in enumerate(geometry_names):
        if geom_name not in geometries_data:
            geometries_data[geom_name] = {
                'available_sizes': [],
                'surface_areas': {}
            }
        
        col = geometry_columns[idx]
        front_area = ws.cell(row=row, column=col).value
        back_area = ws.cell(row=row, column=col + 1).value
        
        # Handle formula cells - evaluate simple formulas
        if hasattr(front_area, 'value'):
            front_area = front_area.value
        if hasattr(back_area, 'value'):
            back_area = back_area.value
        
        # Evaluate simple Excel formulas
        if isinstance(front_area, str) and front_area.startswith('='):
            try:
                # Simple formula evaluation for patterns like "=0.2*0.17"
                front_area = eval(front_area[1:].replace(',', '.'))
            except:
                front_area = 0.0
        
        if isinstance(back_area, str) and back_area.startswith('='):
            try:
                back_area = eval(back_area[1:].replace(',', '.'))
            except:
                back_area = 0.0
        
        # Convert to float if possible
        try:
            front_area = float(front_area) if front_area else 0.0
        except (TypeError, ValueError):
            front_area = 0.0
        
        try:
            back_area = float(back_area) if back_area else 0.0
        except (TypeError, ValueError):
            back_area = 0.0
        
        if size not in geometries_data[geom_name]['surface_areas']:
            geometries_data[geom_name]['surface_areas'][size] = {}
        
        # Only add size if it has actual data (not both zero)
        if front_area > 0 or back_area > 0:
            geometries_data[geom_name]['surface_areas'][size] = {
                'front': front_area,
                'back': back_area
            }
            
            if size not in geometries_data[geom_name]['available_sizes']:
                geometries_data[geom_name]['available_sizes'].append(size)

# Clean up empty surface areas
for geom_name in geometries_data:
    geometries_data[geom_name]['surface_areas'] = {
        size: areas for size, areas in geometries_data[geom_name]['surface_areas'].items()
        if areas  # Remove empty entries
    }
    # Update available sizes to match actual data
    geometries_data[geom_name]['available_sizes'] = sorted(
        geometries_data[geom_name]['surface_areas'].keys()
    )

print(f"\nParsed geometry data:")
for geom_name, data in geometries_data.items():
    print(f"\n{geom_name}:")
    print(f"  Available sizes: {data['available_sizes']}")
    print(f"  Surface areas: {data['surface_areas']}")

# Insert into database
db = SessionLocal()

try:
    for geom_name, data in geometries_data.items():
        # Check if geometry already exists
        existing = db.query(Geometry).filter(Geometry.name == geom_name).first()
        
        if existing:
            print(f"\nGeometry '{geom_name}' already exists. Skipping.")
            continue
        
        # Determine vest type based on geometry name
        vest_type = 'Soft'  # Default
        if 'SAPI' in geom_name.upper():
            vest_type = 'Hard'
        elif 'LATERAL' in geom_name.upper():
            vest_type = 'Hard'
        
        geometry = Geometry(
            name=geom_name,
            description=f"Geometry from Excel calculator - {geom_name}",
            vest_type=vest_type,
            surface_areas=data['surface_areas'],
            available_sizes=data['available_sizes'],
            includes_hard_plates=(vest_type == 'Hard'),
            notes=f"Imported from Excel Superficies sheet"
        )
        
        db.add(geometry)
        print(f"Created geometry: {geom_name}")
    
    db.commit()
    print("\nSuccessfully imported geometries!")
    
except Exception as e:
    db.rollback()
    print(f"\nError importing geometries: {e}")
    import traceback
    traceback.print_exc()
finally:
    db.close()
