"""
Import geometry material configurations from Excel.
This script parses the Excel sheets to extract material requirements and accessories
for each geometry/size combination and creates GeometryMaterialConfig records.
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from app.db.session import SessionLocal
from app.db.models.geometry import Geometry
from app.db.models.geometry_material_config import GeometryMaterialConfig
from app.db.models.material import Material

# Excel file path
EXCEL_FILE = '/Users/josemariabarbeito/PycharmProjects/DeltaDash/Calculadora de Consumos de Chalecos Balistico - DEV-INEX-PROD-SHEET-01.xlsx'

# Mapping from vest sheet names to geometry names
# Only include geometries that actually exist in the database
vest_to_geometry = {
    'ULTRA STOP TCA II': 'DELTA II',
    # Add other mappings as geometries are created in the database
    # 'STOP III': 'STOP III',
    # 'STOP III_B': 'STOP III_B',
    # 'STOP II': 'STOP II',
    # 'MDS_II': 'MDS II',
    # 'MDS_II_LIGHT': 'MDS II LIGHT',
    # 'MDS_III': 'MDS III',
    # 'DEF_III': 'DEF III',
    # 'ULTRA_STOP_III': 'ULTRA STOP III',
    # 'MISS_III': 'MISS III',
    # 'LADY_STOP_III': 'LADY STOP III',
    # 'LADY_STOP_III_B': 'LADY STOP III_B',
    # 'LADY_STOP_III_C': 'LADY STOP III_C',
    # 'LADY_STOP_II': 'LADY STOP II',
    # 'MS_II': 'MS II',
    # 'GEOTEX_COMFORT_III': 'GEOTEX COMFORT III',
    # 'MULTIHIT_II': 'MULTIHIT II',
}

# Material name mapping from Excel to database
material_name_mapping = {
    'Polietileno UAE UD161 1,6x100 m': 'KL 161',
    'Polietileno UD KL230 1,55x120 m': 'KL 230',
    'Polietileno UD Flex KS3000 1,6x150 m': 'FLEX_KS3000',
    'Aramida UD234 1,6x150 m': 'UD 234',
    'Aramida UD245 1,6x150 m': 'UD 245',
    'Poliester No Tejido Geotextil 500g/m² 1,6x50 m': 'GeoTextil 500',
    'Poliester No Tejido Geotextil Blanco 300g/m² 1,5x100 m': 'Geotextil 300',
    'Poliester no Tejido Geotextil 310 g/m² VL930C 1,5x50 m': 'GeoTexile 310',
    'Antitrauma Vitelmat esp.4mm 1x40 m': 'Vitelmat Antitrauma Pad (4mm)',
    'PEAD Natural esp.0,7mm 440x565 mm': 'PEAD 0.7mm',
    'PEAD Natural esp.0,7 480x610 mm': 'PEAD 0.7mm',
    'PEAD Natural esp.0,7mm 500x630 mm': 'PEAD 0.7mm',
    'Honeycomb PE': 'Honeycomb PE',
    'Plate H68': 'Plate H68',
    'Plate H48': 'Plate H48',
    'Placa RB3': 'Placa RB3',
    'Placa RB4': 'Placa RB4',
    'GeoTextile 600': 'GeoTextile 600',
    'FLEX_KH3110A': 'FLEX_KH3110A',
    'SCAT': 'SCAT',
}

# Outer cover/Accessory material mapping
accessory_material_mapping = {
    'Nylon 70D 200 g/m² con TPU Negro 1,5x100 m': None,  # Need to create this material
    'Nylon Rockdura 500D Negro 1,5x100 m': None,  # Need to create this material
    'Tela Spacer Negra Poly 3D Mesh': None,  # Need to create this material
    'Velcro 100 mm Hook': None,  # Need to create this material
    'Velcro 100 mm Loop': None,  # Need to create this material
    'Velcro 50 mm Hook': None,  # Need to create this material
    'Velcro 50 mm Loop': None,  # Need to create this material
    'Velcro 20 mm Hook': None,  # Need to create this material
    'Velcro 20 mm Loop': None,  # Need to create this material
    'Elástico italiano 100 mm': None,  # Need to create this material
}

# Load the workbook
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)

# Get database session
db = SessionLocal()

try:
    created_count = 0
    skipped_count = 0
    
    for sheet_name, geometry_name in vest_to_geometry.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found, skipping")
            skipped_count += 1
            continue
        
        ws = wb[sheet_name]
        
        # Find the geometry
        geometry = db.query(Geometry).filter(Geometry.name == geometry_name).first()
        if not geometry:
            print(f"Geometry {geometry_name} not found in database, skipping")
            skipped_count += 1
            continue
        
        print(f"\nProcessing {geometry_name} from sheet {sheet_name}")
        
        # GeometryMaterialConfig should ONLY contain carrier/accessories (nylon, velcro, etc.)
        # Ballistic materials are stored in Vest → VestLayer
        accessories = []
        
        for row in range(4, 35):
            material_name = ws.cell(row=row, column=1).value
            layer_count = ws.cell(row=row, column=3).value
            part_number = ws.cell(row=row, column=2).value
            
            if not material_name or not isinstance(material_name, str):
                continue
            
            # Check if this is an outer cover/accessory material (Nylon, Velcro, Elástico, Tela)
            is_outer_cover = any(keyword in material_name for keyword in ['Nylon', 'Velcro', 'Elástico', 'Tela', 'Spacer'])
            
            if is_outer_cover:
                # This is an outer cover/accessory material
                # Try to find or create the material
                material = None
                
                # First try to find by exact name
                material = db.query(Material).filter(Material.name == material_name).first()
                
                # If not found, try to find by partial match
                if not material:
                    search_name = material_name.split()[0] if material_name else ''
                    material = db.query(Material).filter(
                        Material.name.ilike(f'%{search_name}%')
                    ).first()
                
                # If still not found, create it
                if not material:
                    # Determine material class based on name
                    material_class = 'accessory'
                    if 'Nylon' in material_name:
                        material_class = 'fabric'
                    elif 'Velcro' in material_name:
                        material_class = 'accessory'
                    elif 'Elástico' in material_name:
                        material_class = 'accessory'
                    elif 'Tela' in material_name:
                        material_class = 'fabric'
                    
                    # Extract areal density if present (e.g., "200 g/m²")
                    areal_density = None
                    if 'g/m²' in material_name:
                        import re
                        match = re.search(r'(\d+)\s*g/m²', material_name)
                        if match:
                            areal_density = float(match.group(1))
                    
                    # Extract roll dimensions if present (e.g., "1,5x100 m")
                    roll_area = None
                    if 'x' in material_name and 'm' in material_name:
                        import re
                        match = re.search(r'([\d,]+)\s*x\s*([\d,]+)\s*m', material_name)
                        if match:
                            width = float(match.group(1).replace(',', '.'))
                            length = float(match.group(2).replace(',', '.'))
                            roll_area = width * length
                    
                    material = Material(
                        name=material_name,
                        material_class=material_class,
                        areal_density_g_m2=areal_density,
                        roll_area_m2=roll_area
                    )
                    db.add(material)
                    db.commit()
                    db.refresh(material)
                    print(f"  Created new material: {material.name}")
                
                # Get quantity from size columns (columns 5-9 for XS-XXL)
                # Use the first size's quantity as default (column 5 = XS)
                quantity = ws.cell(row=row, column=5).value
                
                # If column 5 is a formula, try to evaluate it or use column 3 (layer count)
                if quantity and isinstance(quantity, str) and quantity.startswith('='):
                    # Try to get the value from another column or use layer count
                    quantity = layer_count if layer_count and isinstance(layer_count, (int, float)) else None
                
                if quantity and isinstance(quantity, (int, float)):
                    accessories.append({
                        "material_id": str(material.id),
                        "quantity_per_vest": float(quantity),
                        "unit": "meters",
                        "notes": f"Part: {part_number if part_number else 'N/A'}"
                    })
                    print(f"  Added accessory: {material.name} - {quantity}m per vest")
            # Skip ballistic materials - they belong in Vest → VestLayer
        
        # Create configuration for "ALL" sizes
        if accessories:
            # Check if config already exists
            existing = db.query(GeometryMaterialConfig).filter(
                GeometryMaterialConfig.geometry_id == geometry.id,
                GeometryMaterialConfig.size == "ALL"
            ).first()
            
            if existing:
                print(f"  Config already exists for {geometry_name} - ALL, skipping")
                skipped_count += 1
            else:
                config = GeometryMaterialConfig(
                    geometry_id=geometry.id,
                    size="ALL",
                    material_requirements=[],  # Empty - ballistic materials are in Vest → VestLayer
                    accessories=accessories,
                    efficiency_factor=1.15,  # Default efficiency factor
                    notes=f"Imported from Excel sheet {sheet_name}"
                )
                db.add(config)
                created_count += 1
                print(f"  Created config for {geometry_name} - ALL with {len(accessories)} accessories")
        else:
            print(f"  No accessories found for {geometry_name}")
            skipped_count += 1
    
    db.commit()
    print(f"\nSuccessfully created {created_count} geometry material configurations")
    print(f"Skipped {skipped_count} configurations")
    
except Exception as e:
    db.rollback()
    print(f"\nError importing configurations: {e}")
    import traceback
    traceback.print_exc()
finally:
    db.close()
