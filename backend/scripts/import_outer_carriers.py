"""
Import outer carrier fabric information from Excel.
Outer carrier materials are typically Nylon fabrics for the vest cover.
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from app.db.session import SessionLocal
from app.db.models.geometry import Geometry
from app.db.models.material import Material

# Excel file path
EXCEL_FILE = '/Users/josemariabarbeito/PycharmProjects/DeltaDash/Calculadora de Consumos de Chalecos Balistico - DEV-INEX-PROD-SHEET-01.xlsx'

# Load the workbook
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)

# Mapping from vest sheet names to geometry names
vest_to_geometry = {
    'ULTRA STOP TCA II': 'DELTA II',
    'STOP III': 'STOP III',
    'STOP III_B': 'STOP III_B',
    'STOP II': 'STOP II',
    'MDS_II': 'MDS II',
    'MDS_II_LIGHT': 'MDS II LIGHT',
    'MDS_III': 'MDS III',
    'DEF_III': 'DEF III',
    'ULTRA_STOP_III': 'ULTRA STOP III',
    'MISS_III': 'MISS III',
    'LADY_STOP_III': 'LADY STOP III',
    'LADY_STOP_III_B': 'LADY STOP III_B',
    'LADY_STOP_III_C': 'LADY STOP III_C',
    'LADY_STOP_II': 'LADY STOP II',
    'MS_II': 'MS II',
    'GEOTEX_COMFORT_III': 'GEOTEX COMFORT III',
    'MULTIHIT_II': 'MULTIHIT II',
}

# Outer carrier material from Excel (Nylon 70D 200 g/m² con TPU Negro)
outer_carrier_material_name = 'Nylon 70D 200 g/m² con TPU Negro 1,5x100 m'

# Update geometries with outer carrier info
db = SessionLocal()

try:
    # Find the outer carrier material in the database
    # Try to find a nylon material
    outer_carrier_material = db.query(Material).filter(
        Material.name.ilike('%nylon%')
    ).first()
    
    if not outer_carrier_material:
        print("No nylon material found in database. Please create one first.")
    else:
        print(f"Using outer carrier material: {outer_carrier_material.name}")
    
    updated_count = 0
    for sheet_name, geometry_name in vest_to_geometry.items():
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        # Get outer carrier layer count from row 18, column 3
        layer_count = ws.cell(row=18, column=3).value
        
        if layer_count and outer_carrier_material:
            # Find geometry
            geometry = db.query(Geometry).filter(Geometry.name == geometry_name).first()
            if geometry:
                geometry.outer_carrier_material_id = outer_carrier_material.id
                geometry.outer_carrier_layer_count = int(layer_count)
                print(f"Updated {geometry_name} with outer carrier: {outer_carrier_material.name}, layers: {layer_count}")
                updated_count += 1
    
    db.commit()
    print(f"\nSuccessfully updated {updated_count} geometries with outer carrier information!")
    
except Exception as e:
    db.rollback()
    print(f"\nError updating geometries: {e}")
    import traceback
    traceback.print_exc()
finally:
    db.close()
