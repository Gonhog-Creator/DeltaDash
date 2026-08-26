import re
import uuid
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.db.models.geometry import Geometry


# Excel sheet names cannot contain: []:*?/\
_SHEET_INVALID = re.compile(r"[\\/:?*\[\]]")
MEASUREMENT_KEY_RE = re.compile(r"^(Front|Back)\s+([A-Z])$", re.IGNORECASE)


def _sanitize_sheet_name(name: str) -> str:
    """Sanitize a geometry name so it is a valid Excel sheet name."""
    sanitized = _SHEET_INVALID.sub("_", name)
    # Excel sheet name limit is 31 characters
    if len(sanitized) > 31:
        sanitized = sanitized[:31]
    return sanitized


def _find_detail_sheet(workbook, geometry_name: str):
    """Find a detail sheet matching the sanitized geometry name."""
    sanitized = _sanitize_sheet_name(geometry_name)
    # Try sanitized name first, then exact original (in case it was valid and kept)
    candidates = [sanitized, geometry_name]
    for candidate in candidates:
        if candidate in workbook.sheetnames:
            return workbook[candidate]
    # Fallback: compare case-insensitively
    lower = sanitized.lower()
    for sheet_name in workbook.sheetnames:
        if sheet_name.lower() == lower:
            return workbook[sheet_name]
    return None


def _find_header_row(sheet) -> int:
    """Locate the row containing the 'Size' header in a detail sheet."""
    for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == "size":
                return row_idx
    raise ValueError(f"Could not find 'Size' header in sheet '{sheet.title}'")


def _coerce_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "yes", "1", "y", "si", "sí"}


def _coerce_number(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except (ValueError, TypeError):
        return None


def _build_detail_columns(geometry: Geometry) -> Tuple[List[str], List[str]]:
    """Return the ordered list of front/back measurement keys for a geometry."""
    keys = set()
    measurements = geometry.size_measurements or {}
    for size_data in measurements.values():
        for panel in ("front", "back"):
            panel_data = size_data.get(panel) or {}
            for key in panel_data.keys():
                if isinstance(key, str):
                    keys.add(key)
    sorted_keys = sorted(keys)
    return sorted_keys, [f"Front {k}" for k in sorted_keys] + [f"Back {k}" for k in sorted_keys]


def export_geometries_to_excel(db: Session) -> BytesIO:
    """Export all geometries to an Excel workbook ready for bulk editing."""
    geometries = db.query(Geometry).order_by(Geometry.name).all()

    workbook = Workbook()
    # Remove default sheet
    workbook.remove(workbook.active)

    # Master sheet with geometry metadata
    master = workbook.create_sheet("Geometries")
    master_headers = [
        "name",
        "description",
        "vest_type",
        "includes_hard_plates",
        "is_approved",
        "image_url",
        "compatibility",
        "notes",
    ]
    master.append(master_headers)

    for geometry in geometries:
        master.append([
            geometry.name,
            geometry.description or "",
            geometry.vest_type or "",
            bool(geometry.includes_hard_plates),
            bool(geometry.is_approved),
            geometry.image_url or "",
            geometry.compatibility or "",
            geometry.notes or "",
        ])

    # Per-geometry detail sheets
    for geometry in geometries:
        sheet_name = _sanitize_sheet_name(geometry.name)
        # Ensure unique sheet name
        original_sheet_name = sheet_name
        suffix = 1
        while sheet_name in workbook.sheetnames:
            suffix_str = f"_{suffix}"
            sheet_name = original_sheet_name[: (31 - len(suffix_str))] + suffix_str
            suffix += 1

        sheet = workbook.create_sheet(sheet_name)
        sheet["A1"] = f"Geometry: {geometry.name}"

        measurement_keys, measurement_headers = _build_detail_columns(geometry)
        headers = ["Size", "Front Area", "Back Area", "Total Area"] + measurement_headers
        sheet.append([""] * len(headers))  # row 2 placeholder; will overwrite below
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=2, column=col, value=header)

        available_sizes = geometry.available_sizes or []
        surface_areas = geometry.surface_areas or {}
        size_measurements = geometry.size_measurements or {}

        for size in available_sizes:
            areas = surface_areas.get(size) or {}
            row = [size]
            row.append(_coerce_number(areas.get("front")))
            row.append(_coerce_number(areas.get("back")))
            row.append(_coerce_number(areas.get("total")))

            size_data = size_measurements.get(size) or {}
            front = size_data.get("front") or {}
            back = size_data.get("back") or {}
            for key in measurement_keys:
                row.append(_coerce_number(front.get(key)))
            for key in measurement_keys:
                row.append(_coerce_number(back.get(key)))

            sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _parse_detail_sheet(sheet) -> Tuple[List[str], Dict[str, Dict], Dict[str, Dict]]:
    """Parse a detail sheet and return available sizes, surface_areas, size_measurements."""
    header_row = _find_header_row(sheet)
    headers = [cell.value for cell in sheet[header_row]]

    # Map column names to indices
    col_map = {}
    for idx, header in enumerate(headers, start=1):
        if header is not None:
            col_map[str(header).strip()] = idx

    if "Size" not in col_map:
        raise ValueError(f"Sheet '{sheet.title}' is missing a 'Size' column")

    # Determine measurement columns
    front_keys: Dict[int, str] = {}
    back_keys: Dict[int, str] = {}
    for header, idx in col_map.items():
        match = MEASUREMENT_KEY_RE.match(header)
        if match:
            panel = match.group(1).lower()
            key = match.group(2).upper()
            if panel == "front":
                front_keys[idx] = key
            else:
                back_keys[idx] = key

    available_sizes: List[str] = []
    surface_areas: Dict[str, Dict] = {}
    size_measurements: Dict[str, Dict] = {}

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        size = row[col_map["Size"] - 1] if col_map.get("Size") else None
        if size is None or str(size).strip() == "":
            continue
        size = str(size).strip().upper()
        available_sizes.append(size)

        front_area = _coerce_number(row[col_map.get("Front Area", -1) - 1] if col_map.get("Front Area") else None)
        back_area = _coerce_number(row[col_map.get("Back Area", -1) - 1] if col_map.get("Back Area") else None)
        total_area = _coerce_number(row[col_map.get("Total Area", -1) - 1] if col_map.get("Total Area") else None)

        surface_areas[size] = {}
        if front_area is not None:
            surface_areas[size]["front"] = front_area
        if back_area is not None:
            surface_areas[size]["back"] = back_area
        if total_area is not None:
            surface_areas[size]["total"] = total_area

        front_measurements = {}
        back_measurements = {}
        for col_idx, key in front_keys.items():
            value = _coerce_number(row[col_idx - 1])
            if value is not None:
                front_measurements[key] = value
        for col_idx, key in back_keys.items():
            value = _coerce_number(row[col_idx - 1])
            if value is not None:
                back_measurements[key] = value

        size_measurements[size] = {}
        if front_measurements:
            size_measurements[size]["front"] = front_measurements
        if back_measurements:
            size_measurements[size]["back"] = back_measurements

    return available_sizes, surface_areas, size_measurements


def import_geometries_from_excel(file_path: str, db: Session) -> List[Geometry]:
    """Import or update geometries from an Excel workbook."""
    workbook = load_workbook(file_path, data_only=True)

    if "Geometries" not in workbook.sheetnames:
        raise ValueError("Workbook must contain a 'Geometries' sheet")

    master = workbook["Geometries"]
    master_headers = [cell.value for cell in master[1]]
    col_map = {str(header).strip().lower(): idx for idx, header in enumerate(master_headers, start=1)}

    def get_value(row, header: str):
        idx = col_map.get(header.lower())
        if idx is None:
            return None
        return row[idx - 1]

    upserted: List[Geometry] = []

    for row in master.iter_rows(min_row=2, values_only=True):
        name = get_value(row, "name")
        if not name or str(name).strip() == "":
            continue
        name = str(name).strip()

        detail_sheet = _find_detail_sheet(workbook, name)
        if not detail_sheet:
            raise ValueError(f"Detail sheet not found for geometry '{name}'")

        available_sizes, surface_areas, size_measurements = _parse_detail_sheet(detail_sheet)

        geometry = db.query(Geometry).filter(Geometry.name == name).first()
        if not geometry:
            geometry = Geometry(id=uuid.uuid4())
            db.add(geometry)

        geometry.name = name
        geometry.description = get_value(row, "description") or None
        geometry.vest_type = get_value(row, "vest_type") or None
        geometry.includes_hard_plates = _coerce_bool(get_value(row, "includes_hard_plates"))
        geometry.is_approved = _coerce_bool(get_value(row, "is_approved"))
        geometry.image_url = get_value(row, "image_url") or None
        geometry.compatibility = get_value(row, "compatibility") or None
        geometry.notes = get_value(row, "notes") or None
        geometry.available_sizes = available_sizes
        geometry.surface_areas = surface_areas if surface_areas else None
        geometry.size_measurements = size_measurements if size_measurements else None

        upserted.append(geometry)

    db.commit()
    for geometry in upserted:
        db.refresh(geometry)

    return upserted
